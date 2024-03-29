from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.forms import formset_factory
from django.http import HttpResponse
from .forms import TrainCreationForm, RouteForm, ScheduleForm
from .models import Train, TrainRun, Route, Schedule
from stations.models import Station
from .tasks import schedule
from tickets.models import Ticket
from django.contrib import messages
from datetime import timedelta, datetime
from django.utils import timezone
from openpyxl import Workbook
from django.core.paginator import Paginator

@login_required(login_url='accounts/login/')
def CreateTrain(request):
    if request.user.is_staff:
        message = None
        if request.method == 'POST':
            form = TrainCreationForm(request.POST)
            if form.is_valid():
                num_stops = form.cleaned_data.get('numberOfStops')
                train = form.save(commit=False)
                train.save()
                train.daysOfWeek.set(form.cleaned_data.get('days_of_week'))
                message = messages.success(request, 'Train successfully created')
                create_train_runs(train)

                return redirect('create-train-route', pk=train.id, num_stops=num_stops)
            else:
                message = messages.error(request, f'{form.errors}')
        form = TrainCreationForm()
        context = {'form': form, 'messsage': message}
        return render(request, 'trains/create_train.html', context)
    
    else:
        message = messages.warning(request, "You don't have the permission to visit this page!")
        return redirect('home')


def create_train_runs(train):
        current_date = timezone.now().date()
        end_date = current_date + timedelta(days=30)

        train_days = set(train.daysOfWeek.values_list('day_code', flat=True))

        while current_date <= end_date:
            if current_date.strftime('%a') in train_days:
                TrainRun.objects.create(
                    train=train,
                    departure_date=current_date,
                    arrival_date=current_date + timedelta(days=train.daysOfJourney),
                    numberOfAvailable1AC=train.numberOf1AC,
                    numberOfAvailable2AC=train.numberOf2AC,
                    numberOfAvailable3AC=train.numberOf3AC,
                    numberOfAvailableSleeper=train.numberOfSleeper,
                )

            current_date += timedelta(days=1)

def TrainScheduleView(request, pk, num_stops):
    train = get_object_or_404(Train, id=pk)
    message = None
    routes = Route.objects.filter(train=train)
    stations = []
    for route in routes:
        station = route.station
        stations.append(station)

    ScheduleFormSet = formset_factory(ScheduleForm, extra=num_stops)
    if request.method == 'POST':
        schedule_formset = ScheduleFormSet(request.POST)
        trainRuns = TrainRun.objects.filter(train=train)
        if schedule_formset.is_valid():
            for trainRun in trainRuns:
                i=1
                schedules = []
                days = []
                for schedule_form in schedule_formset:
                    daysRequiredToReach = schedule_form.cleaned_data.get('daysRequiredToReach')
                    arrivalTime = schedule_form.cleaned_data.get('arrivalTime')
                    departureTime = schedule_form.cleaned_data.get('departureTime')
                    station = stations[i]
                    date = trainRun.departure_date + timedelta(days=daysRequiredToReach)
                    date_time = datetime.datetime()
                    if date <= trainRun.arrival_date:
                        schedule = Schedule.objects.create(trainRun=trainRun, station=station, date=date, daysRequiredToReach=daysRequiredToReach, arrivalTime=arrivalTime, departureTime=departureTime)
                        schedules.append(schedule)
                        if i != len(stations)-2:
                            i+=1
                    else:
                        messages.error(request, "No intermediate stops can have days required to reach greater than total days of train's journey!")
                        for schedule in schedule:
                            schedule.delete()
                        return redirect('create-train-schedule', pk=pk, num_stops=num_stops)
                Schedule.objects.create(trainRun=trainRun, station=train.source, daysRequiredToReach=0, date=trainRun.departure_date, arrivalTime=train.departureTime, departureTime=train.departureTime)
                Schedule.objects.create(trainRun=trainRun, station=train.destination, daysRequiredToReach=train.daysOfJourney, date=trainRun.arrival_date, arrivalTime=train.arrivalTime, departureTime=train.arrivalTime)
        else:
            train.delete()
            print(schedule_formset.errors)
            message = messages.error(request, 'Error validating form!')
        
        return redirect('all-trains')
    else:
        schedule_formset = ScheduleFormSet()
        dict = {}
        i=0
        for schedule_form in schedule_formset:
            dict[schedule_form] = stations[i]
            i+=1
    context = {'dict': dict.items(), 'schedule_formset': schedule_formset, 'message': message}
    return render(request, 'trains/create_train_schedule.html', context)

def TrainRouteView(request, pk, num_stops):
    train = get_object_or_404(Train, id=pk)
    message = None
    RouteFormSet = formset_factory(RouteForm, extra=num_stops)
    if request.method == 'POST':
        route_formset = RouteFormSet(request.POST)
        for route_form in route_formset:
            if route_form.is_valid():
                if route_form.cleaned_data.get('station') != train.source and route_form.cleaned_data.get('station') != train.destination:
                    if route_form.cleaned_data.get('distance')<train.totalDistance:
                        route = route_form.save(commit=False)
                        route.train = train
                        route.save()
                    else:
                        messages.error(request, "The distance of any intermediate stop should not be greated than train's total distance!")
                        return redirect('create-train-route' ,pk=train.id, num_stops=num_stops)
                else:
                    messages.error(request, "The train cannot have a stop at the source or destination!")
                    return redirect('create-train-route' ,pk=train.id, num_stops=num_stops)
            else:
                train.delete()
                message = messages.error(request, f"{route_formset.errors}")

        Route.objects.create(train=train, station=train.destination, distance=train.totalDistance)
        Route.objects.create(train=train, station=train.source, distance=0)
        return redirect('create-train-schedule', pk=train.id, num_stops=num_stops)
    else:
        route_formset = RouteFormSet()
    context = {'formset': route_formset, 'message': message}
    return render(request, 'trains/create_train_route.html', context)

@login_required(login_url='accounts/login/')
def ViewTrainRoute(request, pk):
    train = Train.objects.get(id=pk)
    train_route = Route.objects.filter(train=train).order_by('distance')

    context = {'train': train, 'train_route': train_route}
    return render(request, 'trains/view_train_route.html', context)

@login_required(login_url='accounts/login/')
def AllTicketsTrainView(request, pk):
    user = request.user
    if user.is_staff:
        train = Train.objects.get(id=pk)
        trainRuns = TrainRun.objects.filter(train=train)
        all_tickets = []
        for trainRun in trainRuns:
            tickets = Ticket.objects.filter(trainRun=trainRun)
            for ticket in tickets:
                if ticket.date>datetime.today().date():
                    all_tickets.append(ticket)
        
        context = {'all_tickets': all_tickets, 'train': train, 'type': 'train'}
        return render(request, 'trains/all_tickets.html', context)

    else:
        messages.warning(request, "You are not allowed to visit that page!!!")
        return redirect('home')

@login_required(login_url='accounts/login/')
def AllTicketsTrainRunView(request, pk):
    user = request.user
    if user.is_staff:
        trainRun = TrainRun.objects.get(id=pk)
        tickets = Ticket.objects.filter(trainRun=trainRun)
        all_tickets = []
        for ticket in tickets:
            if ticket.date>datetime.today().date():
                all_tickets.append(ticket)
        
        context = {'all_tickets': all_tickets, 'train': trainRun, 'type': 'trainRun'}
        return render(request, 'trains/all_tickets.html', context)

    else:
        messages.warning(request, "You are not allowed to visit that page!!!")
        return redirect('home')
    

@login_required(login_url='accounts/login/')
def AllTrainsView(request):
    user = request.user
    if user.is_staff:
        trains = Train.objects.all()
        context = {'trains': trains}
        return render(request, 'trains/all_trains.html', context)
    
    else:
        messages.warning(request, "You are not allowed to visit that page!!!")
        return redirect('home')
    
@login_required(login_url='accounts/login/')
def AllTrainRunsView(request, pk):
    user = request.user
    if user.is_staff:
        train = Train.objects.get(id=pk)
        trainRuns = TrainRun.objects.filter(train=train)
        context = {'trainRuns': trainRuns, 'train': train}
        return render(request, 'trains/all_trainRuns.html', context)
    
    else:
        messages.warning(request, "You are not allowed to visit that page!!!")
        return redirect('home')
    
@login_required(login_url='accounts/login/')
def EditTrianView(request, pk):
    user=request.user
    if user.is_staff:
        message = None
        train = Train.objects.get(id=pk)
        if request.method == 'POST':
            form = TrainCreationForm(request.POST, instance=train)
            if form.is_valid():
                updated_train = form.save(commit=False)
                updated_train.id=pk
                updated_train.save()
                message = messages.success(request, "The Train details have been updated successfully.")
                return redirect('all-trains')
            else:
                message = messages.error(request, "Error validating Form!!")
        else:
            form = TrainCreationForm(instance=train)
        context = {'form': form, 'message': message}
        return render(request, 'trains/create_train.html', context)
    else:
        message = messages.warning(request, "You are not allowed to visit that page!!!")
        return redirect('home')

@login_required(login_url='accounts/login/')
def DeleteTrainView(request, pk):
    user = request.user
    if user.is_staff:
        train = Train.objects.get(id=pk)
        if request.method == 'POST':
            trainRuns = TrainRun.objects.filter(train=train)
            tickets = []
            for trainRun in trainRuns:
                ticket_trainRun = Ticket.objects.filter(trainRun = trainRun)
                tickets.append(ticket_trainRun)
            
            for ticket_trainRun in tickets:
                for ticket in ticket_trainRun:
                    ticket_user = ticket.user
                    profile = ticket_user.profile
                    profile.wallet += ticket.fare
                    profile.save()
            train.delete()

            messages.success(request, "Train deleted successfully.")
            return redirect('all-trains')
        
        return render(request, 'delete.html', {'obj': train})
    
    else:
        messages.warning(request, "You are not allowed to visit that page!!!")
        return redirect('home')
    
@login_required(login_url='accounts/login/')
def DeleteTrainRunView(request, pk):
    user = request.user
    if user.is_staff:
        trainRun = TrainRun.objects.get(id=pk)
        if request.method == 'POST':
            tickets = Ticket.objects.filter(trainRun = trainRun)
            
            for ticket in tickets:
                ticket_user = ticket.user
                profile = ticket_user.profile
                profile.wallet += ticket.fare
                profile.save()
            trainRun.delete()

            messages.success(request, "TrainRun deleted successfully.")
            return redirect('all-trainruns')
        
        return render(request, 'delete.html', {'obj': trainRun})
    
    else:
        messages.warning(request, "You are not allowed to visit that page!!!")
        return redirect('home')

@login_required(login_url='accounts/login/')
def ExportFile(request):
    today = datetime.today().date()
    pk = request.GET.get('pk')
    type = request.GET.get('type')
    if type=='train':
        train = Train.objects.get(id=pk)
        trainRuns = TrainRun.objects.filter(train=train).order_by('departure_date')

        all_tickets = []
        for trainRun in trainRuns:
            tickets = Ticket.objects.filter(trainRun=trainRun)
            for ticket in tickets:
                if ticket.date > today:
                    all_tickets.append(ticket)
    
    else:
        trainRun = TrainRun.objects.get(id=pk)
        train = TrainRun.train
        all_tickets = []
        tickets = Ticket.objects.filter(trainRun=trainRun)
        for ticket in tickets:
            if ticket.date > today:
                all_tickets.append(ticket)        

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f"attachment; filename={train}_tickets.xlsx"
    
    workbook = Workbook()

    worksheet = workbook.active

    worksheet.merge_cells('A1:J1')
    first_cell = worksheet['A1']
    first_cell.value = f"tickets for {train}"

    columns = ['user', 'date', 'departure_station', 'destination_station', 'booking_time', 'passenger', 'seatNumber', 'seatClass', 'status', 'fare']
    row_num = 2

    for col_num, col_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = col_title
        
    for ticket in all_tickets:
        row_num+=1
        date = ticket.date.strftime("%Y-%m-%d")
        booking_time = ticket.booking_time.strftime("%Y-%m-%d %H:%M:%S")

        row = [ticket.user.username, date, ticket.departure_station.stationName, ticket.destination_station.stationName, booking_time, ticket.passenger.name, ticket.seatNumber, ticket.seatClass.seat_class, ticket.status, ticket.fare]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response

def daily_trainruns(request):
    schedule.delay()
    return HttpResponse("New Train runs have been created successfully!!")